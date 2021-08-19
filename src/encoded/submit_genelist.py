import json
import re
from base64 import b64encode

from dcicutils.misc_utils import VirtualApp, VirtualAppError
from openpyxl import load_workbook
from webtest import AppError

from .util import s3_local_file
from .ingestion.common import CGAP_CORE_PROJECT

CGAP_CORE_PROJECT = CGAP_CORE_PROJECT + "/"


def submit_genelist(
    *, s3_client, bucket, key, project, institution, vapp, validate_only=False
):
    """
    Handles processing of a submitted gene list.

    :param s3_client: a boto3 s3 client object
    :param bucket: str name of the s3 bucket that contains the data to be
        processed
    :param key: str name of key within the given bucket that contains the data
        to be processed
    :param project: str project identifier
    :param institution: str institution identifier
    :param vapp: a vapp object
    :param validate_only: bool if True, only do validation, not posting;
        otherwise, do posting, too.

    :returns: dict information concerning output of processing
    """

    with s3_local_file(s3_client, bucket=bucket, key=key) as filename:
        results = {
            "success": False,
            "validation_output": [],
            "result": {},
            "post_output": [],
            "upload_info": [],
        }
        if filename.endswith((".txt", ".xlsx")):
            genelist = GeneListSubmission(
                filename, project, institution, vapp, validate_only
            )
            if validate_only:
                if genelist.errors:
                    results["validation_output"] = genelist.errors
                else:
                    results["success"] = True
                    results["validation_output"] = (
                        genelist.validation_output + genelist.notes
                    )
            elif genelist.post_output:
                results["success"] = True
                results["validation_output"] = (
                    genelist.validation_output + genelist.notes
                )
                results["post_output"] = genelist.post_output
                results["result"] = {"genelist": genelist.at_id}
            else:
                results["validation_output"] = genelist.errors
        else:
            msg = "Gene list must be a .txt or .xlsx file."
            results["validation_output"].append(msg)
        return results


class GeneListSubmission:
    """
    Class to handle processing of submitted gene list.
    """

    def __init__(self, filename, project, institution, vapp, validate_only=False):
        self.filename = filename
        self.project = project
        self.institution = institution
        self.vapp = vapp
        self.errors = []
        self.notes = []
        self.non_gene_terms = ["title", "case"]
        self.genes, self.title, self.case_atids = self.parse_genelist()
        self.gene_ids = self.match_genes()
        self.bam_sample_ids = self.get_case_sample_id()
        self.post_bodies = self.create_post_bodies()
        (
            self.validation_output,
            self.patch_genelist_uuid,
            self.patch_document_uuid,
            self.previous_gene_ids,
            self.previous_bam_ids,
        ) = self.validate_postings()
        if not validate_only:
            self.post_output, self.at_id = self.post_items()
            self.submit_variant_update()

    def extract_txt_file(self, txt_file):
        """
        Parses a .txt file for genes, cases, and title. For every non-empty
        line, searches for starting word of "title" or "case" and makes
        corresponding assignment; otherwise, line is assumed to be a gene.

        :param txt_file: file object
        :returns:
            - title - str gene list title
            - cases - list of strings of associated cases
            - genelist - list of genes from provided gene list
        """
        genelist = []
        non_gene_items = {term: [] for term in self.non_gene_terms}
        with open(txt_file, "r", encoding="utf-8") as genelist_file:
            genelist_raw = genelist_file.readlines()
        for line in genelist_raw.copy():
            for term in self.non_gene_terms:
                if line.lower().startswith(term):
                    term_line = line
                    term_idx = term_line.lower().index(term)
                    term_line = term_line[(term_idx + len(term)) :]
                    term_line = term_line.translate({ord(i): None for i in ':;"\n"'})
                    term_line = term_line.strip()
                    if term_line:
                        non_gene_items[term].append(term_line)
                    genelist_raw.remove(line)
                    break
        title = non_gene_items["title"]
        cases = non_gene_items["case"]
        for line in genelist_raw:
            line = line.translate({ord(i): None for i in '"\t""\n":;'})
            genelist.append(line.strip())
        return title, cases, genelist

    def extract_xlsx_file(self, xlsx_file):
        """
        Parses a .xlsx file for genes, cases, and title. Assumes headers "Genes",
        "Case", and "Title" are in the first row of document and will designate
        the rows underneath appropriately (though title is taken to be the first
        non-empty row under the header).

        :param xlsx_file: file object
        :returns:
            - title - str gene list title
            - cases - list of strings of associated cases
            - genelist - list of genes from provided gene list
        """
        possible_items = {term: [] for term in self.non_gene_terms}
        possible_items["genes"] = []
        wb = load_workbook(xlsx_file)
        sheet = wb.worksheets[0]
        for col in sheet.iter_cols(values_only=True):
            if isinstance(col[0], str) and len(col) > 1:
                for term in possible_items:
                    if term in col[0].lower():
                        for cell in col[1:]:
                            if not cell:
                                continue
                            elif isinstance(cell, str):
                                possible_items[term].append(cell.strip())
                            elif isinstance(cell, int):
                                possible_items[term].append(str(cell))
                        break
        title = possible_items["title"]
        cases = possible_items["case"]
        genelist = possible_items["genes"]
        return title, cases, genelist

    @staticmethod
    def _create_case_atids(case_list):
        """
        Convert case accession list to list of case @ids.

        :param case_list: list of strings of case identifiers
        :returns: list of strings of potential case @ids
        """
        cases = []
        case_atids = []
        for case in case_list:
            if "," in case:
                cases += case.split(",")
            else:
                cases.append(case)
        cases = [case.strip() for case in cases if case]
        cases = list(set(cases))
        case_atids = ["/cases/" + case + "/" for case in cases]
        return case_atids

    def parse_genelist(self):
        """
        Parses gene list file, extracts title and cases, and removes any duplicate
        genes.

        Assumes gene list is txt file with line-separated title/genes or an
        xlsx file with standard formatting defined.

        :returns:
            - genelist - list of strings of unique genes as provided
                in submission
            - title - str gene list title (or None)
            - cases - list of strings of case @ids
        """

        genes_with_spaces = []
        non_ascii_genes = []
        if self.filename.endswith(".txt"):
            title, cases, genelist = self.extract_txt_file(self.filename)
        elif self.filename.endswith(".xlsx"):
            title, cases, genelist = self.extract_xlsx_file(self.filename)
        if title and title[0]:
            title = title[0].translate({ord(i): None for i in "'+&!?=%/"}).strip()
        else:
            title = None
            self.errors.append(
                "No title was found in the gene list. Please check the "
                "formatting of the submitted document."
            )
        case_atids = self._create_case_atids(cases)
        genelist = [x for x in genelist if x != ""]
        for gene in genelist.copy():
            if " " in gene:
                genelist.remove(gene)
                genes_with_spaces.append(gene)
            if not CommonUtils.is_ascii(gene):
                genelist.remove(gene)
                non_ascii_genes.append(gene)
        if genes_with_spaces:
            self.errors.append(
                "Gene symbols/IDs should not contain spaces. Please reformat "
                "the following gene entries: %s." % ", ".join(genes_with_spaces)
            )
        if non_ascii_genes:
            self.errors.append(
                "The following gene(s) contain non-ASCII characters: %s. "
                "Please re-enter the genes using only ASCII characters."
                % ", ".join(non_ascii_genes)
            )
        if not genelist and not genes_with_spaces and not non_ascii_genes:
            self.errors.append(
                "No genes were found in the gene list. Please check the "
                "formatting of the submitted document."
            )
        genelist = list(set(genelist))
        return genelist, title, case_atids

    def match_genes(self):
        """
        Attempts to match every unique gene to unique gene item in CGAP.

        Matches by gene symbol, alias symbol, previous symbol, genereviews
        symbol, OMIM ID, Entrez ID, and UniProt ID, in that order.
        If an Ensembl ID is given, searches for a direct match, and if not
        found no additional search done given strange query behavior.
        If multiple genes in the given gene list direct to the same gene item,
        only one instance is included, so the gene list produced may be shorter
        than the one submitted.

        :returns: list of gene uuids in CGAP gene title alphabetical
            order or None
        """

        if not self.genes:
            return None
        ensgids = []
        non_ensgids = []
        gene_ids = {}
        gene_ensgids = {}
        unmatched_genes_without_options = []
        for gene in self.genes:
            if re.fullmatch(r"ENSG\d{11}", gene):
                ensgids.append(gene)
            else:
                non_ensgids.append(gene)
        if ensgids:
            ensgid_search = CommonUtils.batch_search(
                self.vapp, ensgids, "ensgid", batch_size=10
            )
            for response in ensgid_search:
                if response["gene_symbol"] in gene_ids:
                    gene_ids[response["gene_symbol"]].append(response["uuid"])
                    gene_ensgids[response["gene_symbol"]].append(response["ensgid"])
                else:
                    gene_ids[response["gene_symbol"]] = [response["uuid"]]
                    gene_ensgids[response["gene_symbol"]] = [response["ensgid"]]
                ensgids.remove(response["ensgid"])
            if ensgids:
                unmatched_genes_without_options += ensgids
        if non_ensgids:
            search_order = [
                "gene_symbol",
                "alias_symbol",
                "prev_symbol",
                "genereviews",
                "omim_id",
                "entrez_id",
                "uniprot_ids",
            ]
            for search_type in search_order:
                if not non_ensgids:
                    break
                search = CommonUtils.batch_search(
                    self.vapp,
                    non_ensgids,
                    search_type,
                    batch_size=10,
                )
                for response in search:
                    if (
                        response["gene_symbol"] in gene_ids
                        and response["uuid"] not in gene_ids[response["gene_symbol"]]
                    ):
                        gene_ids[response["gene_symbol"]].append(response["uuid"])
                        gene_ensgids[response["gene_symbol"]].append(response["ensgid"])
                        responsible_gene = response[search_type]
                        if type(response[search_type]) is list:
                            for item in response[search_type]:
                                if item in gene_ensgids:
                                    responsible_gene = item
                                    break
                        self.notes.append(
                            "Note: gene %s refers to multiple genes in our database,"
                            " including genes with the following Ensembl IDs: %s."
                            " If you would prefer to only include one of these genes"
                            " in this gene list, please resubmit and replace the"
                            " gene %s with one of the Ensembl IDs above."
                            % (
                                responsible_gene,
                                ", ".join(gene_ensgids[responsible_gene]),
                                responsible_gene,
                            )
                        )
                        continue
                    else:
                        gene_ids[response["gene_symbol"]] = [response["uuid"]]
                        gene_ensgids[response["gene_symbol"]] = [response["ensgid"]]
                    if type(response[search_type]) is str:
                        non_ensgids.remove(response[search_type])
                    elif type(response[search_type]) is list:
                        for item in response[search_type]:
                            if item in non_ensgids:
                                non_ensgids.remove(item)
                                break
        if non_ensgids:
            for gene in non_ensgids:
                try:
                    response = self.vapp.get("/search/?type=Gene&q=" + gene).json[
                        "@graph"
                    ]
                    options = [option["gene_symbol"] for option in response]
                    self.errors.append(
                        "No perfect match found for gene %s. "
                        "Consider replacing with one of the following: %s."
                        % (gene, ", ".join(options))
                    )
                except (VirtualAppError, AppError):
                    unmatched_genes_without_options.append(gene)
        if unmatched_genes_without_options:
            self.errors.append(
                "The gene(s) %s could not be found in our database. "
                "Consider replacement with an alias name or an Ensembl ID."
                % ", ".join(unmatched_genes_without_options)
            )
        sorted_gene_names = sorted(gene_ids)
        matched_gene_uuids = []
        for gene_name in sorted_gene_names:
            matched_gene_uuids += gene_ids[gene_name]
        return matched_gene_uuids

    @staticmethod
    def _accession_from_atid(atid):
        """Retrieves case/file accessions from @id."""
        return atid.split("/")[-2]

    def get_case_sample_id(self):
        """
        For each case accession provided, identify the associated sample's BAM
        Sample ID, if it exists.

        :returns: list of BAM sample IDs
        """
        bam_sample_ids = []
        project_mismatch = []
        cases_without_sample_ids = []
        cases_not_found = []
        for case_atid in self.case_atids:
            try:
                response = self.vapp.get(case_atid, status=200).json
                case_project = response.get("project").get("@id")
                if self.project != CGAP_CORE_PROJECT and self.project != case_project:
                    project_mismatch.append(self._accession_from_atid(case_atid))
                    continue
                sample = response.get("sample", {})
                case_bam_sample_id = sample.get("bam_sample_id", "")
                if not case_bam_sample_id:
                    cases_without_sample_ids.append(
                        self._accession_from_atid(case_atid)
                    )
                else:
                    bam_sample_ids.append(case_bam_sample_id)
            except (VirtualAppError, AppError):
                cases_not_found.append(self._accession_from_atid(case_atid))
        bam_sample_ids = list(set(bam_sample_ids))
        if project_mismatch:
            self.errors.append(
                "The following cases belonged to projects different than the project"
                " used for gene list submission: %s. Please remove these cases from"
                " this submission or re-submit under the appropriate project."
                % ", ".join(project_mismatch)
            )
        if cases_without_sample_ids:
            self.errors.append(
                "The following cases are missing sample metadta: %s. Please resubmit"
                " the gene list once sample metadata is complete for these cases"
                " or remove these cases from the current submission."
                % ", ".join(cases_without_sample_ids)
            )
        if cases_not_found:
            self.errors.append(
                "No cases could be found that matched the following accessions: %s."
                % ", ".join(cases_not_found)
            )
        return bam_sample_ids

    def create_post_bodies(self):
        """
        Creates gene list and document bodies for posting.

        :returns: list of bodies to post (document, gene list) or None
        """

        if not self.gene_ids:
            return None
        with open(self.filename, "rb") as stream:
            extension = self.filename.split(".")[-1]
            if not self.title:
                self.title = "Stand_in_genelist"
            if extension == "txt":
                content_type = "text/plain"
            elif extension == "xlsx":
                content_type = (
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            attach = {
                "download": self.title.replace(" ", "_") + "_genelist." + extension,
                "type": content_type,
                "href": (
                    "data:%s;base64,%s"
                    % (
                        content_type,
                        b64encode(stream.read()).decode("ascii"),
                    )
                ),
            }
        document_post_body = {
            "institution": self.institution,
            "project": self.project,
            "attachment": attach,
        }
        genelist_post_body = {
            "title": (
                self.title + " (%s)" % len(self.gene_ids)
                if self.title
                else "Stand in title"
            ),
            "institution": self.institution,
            "project": self.project,
            "genes": self.gene_ids,
        }
        if self.bam_sample_ids:
            genelist_post_body["bam_sample_ids"] = self.bam_sample_ids
        return [document_post_body, genelist_post_body]

    def validate_postings(self):
        """
        Attempts to validate document and gene list jsons provided by
        post_bodies.

        Gene lists of the same title belonging to the same project and
        institution will be searched for, and if a match is discovered, the
        gene list and document will be patched and updated. If no match is
        found, a new document and gene list will be posted.

        :returns:
            - validate_result - list with validation information or None
            - genelist_uuid - str previously existing gene list uuid or None
            - document_uuid - str previously existing document uuid or None
            - previous_genelist_gene_ids - list of genes from existing gene
                list or None
            - prior_sample_ids - list of BAM sample IDs from existing gene
                list or None
        """

        if not self.gene_ids:
            return None, None, None, None, None
        validate_result = {}
        validate_display = []
        genelist_uuid = None
        document_uuid = None
        previous_genelist_gene_ids = []
        prior_sample_ids = []
        removed_sample_ids = []
        project_list = [self.project.replace("/", "%2F")]
        document_json = self.post_bodies[0]
        genelist_json = self.post_bodies[1]
        if self.title:
            try:
                project_genelists = CommonUtils.batch_search(
                    self.vapp,
                    project_list,
                    "project.%40id",
                    item_type="GeneList",
                    institution=self.institution,
                )
                project_titles = [x["title"] for x in project_genelists]
                for previous_title in project_titles:
                    genelist_title = previous_title
                    title_count = re.findall(r"\(\d{1,}\)", previous_title)
                    if title_count:
                        title_count_idx = previous_title.index(title_count[-1])
                        previous_title = previous_title[:title_count_idx].strip()
                    if self.title == previous_title:
                        genelist_idx = project_titles.index(genelist_title)
                        old_genelist = project_genelists[genelist_idx]
                        genelist_uuid = old_genelist["uuid"]
                        prior_genes = old_genelist["genes"]
                        prior_sample_ids = old_genelist.get("bam_sample_ids", [])
                        previous_genelist_gene_ids += [
                            gene["uuid"] for gene in prior_genes
                        ]
                        break
            except VirtualAppError:
                pass
        try:
            fields = ["display_title", "uuid"]
            project_documents = CommonUtils.batch_search(
                self.vapp,
                project_list,
                "project.%40id",
                item_type="Document",
                institution=self.institution,
                fields=fields,
            )
            project_docs = [
                x["display_title"].replace(".txt", "").replace(".xlsx", "")
                for x in project_documents
            ]
            doc_title = self.title.replace(" ", "_") + "_genelist"
            if doc_title in project_docs:
                document_idx = project_docs.index(doc_title)
                document_uuid = project_documents[document_idx]["uuid"]
        except VirtualAppError:
            pass
        if document_uuid:
            for key in document_json.copy():
                if key != "attachment":
                    del document_json[key]
            self.post_bodies[0] = document_json
            self.vapp.patch_json(
                "/Document/" + document_uuid + "?check_only=true",
                document_json,
            )
        else:
            self.vapp.post_json("/Document/?check_only=true", document_json)
        validate_result["Document"] = "validated"
        if genelist_uuid:
            for key in ["project", "institution"]:
                del genelist_json[key]
            if prior_sample_ids:
                genelist_json["bam_sample_ids"] = self.bam_sample_ids
            self.post_bodies[1] = genelist_json
            self.vapp.patch_json(
                "/GeneList/" + genelist_uuid + "?check_only=true",
                genelist_json,
            )
        else:
            self.vapp.post_json("/GeneList/?check_only=true", genelist_json)
        validate_result["Gene list"] = "validated"
        validate_result["Title"] = self.title
        validate_result["Number of genes"] = len(genelist_json["genes"])
        if self.bam_sample_ids:
            case_accessions = [
                self._accession_from_atid(case_id) for case_id in self.case_atids
            ]
            validate_result["Associated cases"] = ", ".join(case_accessions)
        if prior_sample_ids:
            removed_sample_ids = [
                sample_id
                for sample_id in prior_sample_ids
                if sample_id not in self.bam_sample_ids
            ]
        if removed_sample_ids:
            case_search = CommonUtils.batch_search(
                self.vapp,
                removed_sample_ids,
                "sample.bam_sample_id",
                item_type="Case",
                fields=["accession"],
            )
            cases_removed = [case["accession"] for case in case_search]
            validate_result["Removed cases"] = ", ".join(cases_removed)
        for key in validate_result:
            validate_display.append("%s: %s" % (key, validate_result[key]))
        if genelist_uuid:
            validate_display.append(
                "Existing gene list with the same title was found and will be "
                "overwritten."
            )
        validate_result = validate_display
        return (
            validate_result,
            genelist_uuid,
            document_uuid,
            previous_genelist_gene_ids,
            prior_sample_ids,
        )

    def post_items(self):
        """
        Attempts to post document and gene list.

        :returns:
            - post_result - list with posting information or None
            - genelist_at_id - str @id of gene list posted/patched
        """
        if self.errors:
            return None, None
        post_result = {}
        document_json = self.post_bodies[0]
        genelist_json = self.post_bodies[1]
        if self.patch_document_uuid:
            document_post = self.vapp.patch_json(
                "/Document/" + self.patch_document_uuid,
                document_json,
            ).json
        else:
            document_post = self.vapp.post_json("/Document/", document_json).json
        post_result["Document"] = (
            "posted with uuid " + document_post["@graph"][0]["uuid"]
        )
        if document_post["status"] == "success":
            genelist_json["source_file"] = document_post["@graph"][0]["@id"]
            if self.patch_genelist_uuid:
                genelist_post = self.vapp.patch_json(
                    "/GeneList/" + self.patch_genelist_uuid, genelist_json
                ).json
            else:
                genelist_post = self.vapp.post_json("/GeneList/", genelist_json).json
            post_result["Gene list"] = (
                "posted with uuid " + genelist_post["@graph"][0]["uuid"]
            )
            genelist_at_id = genelist_post["@graph"][0]["@id"]
        post_display = []
        for key in post_result:
            post_display.append("%s: %s" % (key, post_result[key]))
        post_result = post_display
        return post_result, genelist_at_id

    def submit_variant_update(self):
        """
        If gene list successfully posted, submit the matched gene uuids to the
        'variant_update' endpoint so they can be updated to reflect the new
        gene list.

        Updates self.post_output with message regarding success or failure.
        """
        if not self.post_output:
            return
        bam_samples = []
        uuids_to_update = list(set(self.gene_ids + self.previous_gene_ids))
        if self.bam_sample_ids and self.previous_bam_ids:
            bam_samples = list(set(self.bam_sample_ids + self.previous_bam_ids))
        elif self.bam_sample_ids and not self.patch_genelist_uuid:
            bam_samples = self.bam_sample_ids
        datafile = json.dumps(
            {"gene_uuids": uuids_to_update, "bam_sample_ids": bam_samples}
        )
        creation_post_url = "/IngestionSubmission"
        creation_post_data = {
            "ingestion_type": "variant_update",
            "project": self.project,
            "institution": self.institution,
            "processing_status": {"state": "submitted"},
        }
        creation_response = self.vapp.post_json(
            creation_post_url,
            creation_post_data,
            content_type="application/json",
        ).json
        submission_id = creation_response["@graph"][0]["@id"]
        submission_post_url = submission_id + "submit_for_ingestion"
        submission_post_data = {"validate_only": False}
        upload_file = [
            (
                "datafile",
                self.title.replace(" ", "_") + ".json",
                bytes(datafile, encoding="utf-8"),
            )
        ]
        submission_response = self.vapp.post(
            submission_post_url,
            submission_post_data,
            upload_files=upload_file,
            content_type="multipart/form-data",
        ).json
        if submission_response["success"]:
            self.post_output.append(
                "Variants should begin updating shortly but may take a few "
                "hours depending on server load."
            )
        else:
            self.post_output.append(
                "Variants were not queued for updating. Please reach out "
                "to the CGAP team and alert them of this issue."
            )
        return


def submit_variant_update(
    *, s3_client, bucket, key, project, institution, vapp, validate_only=False
):
    """
    Processes a submitted list of gene uuids to re-index associated variant
    samples.

    :param s3_client: a boto3 s3 client object
    :param bucket: str name of the s3 bucket that contains the data to be
        processed
    :param key: str name of key within the given bucket that contains the data
        to be processed
    :param project: str project identifier
    :param institution: str institution identifier
    :param vapp: a vapp object
    :param validate_only: bool if True, only do validation, not posting;
        otherwise, do posting, too.

    :returns: dict information concerning output of processing
    """

    with s3_local_file(s3_client, bucket=bucket, key=key) as filename:
        results = {
            "success": False,
            "validation_output": [],
            "result": {},
            "post_output": [],
            "upload_info": [],
        }
        if not filename.endswith(".json"):
            msg = "Expected input file to be a json file."
            results["validation_output"].append(msg)
            return results
        variant_update = VariantUpdateSubmission(
            filename, project, institution, vapp, validate_only
        )
        if validate_only:
            results["validation_output"] = (
                [variant_update.validate_output] + variant_update.errors
                if variant_update.errors
                else variant_update.validate_output
            )
        elif variant_update.post_output:
            results["success"] = True
            results["validation_output"] = variant_update.validate_output
            results["post_output"] = variant_update.post_output
        else:
            results["validation_output"] = variant_update.errors + [
                variant_update.validate_output
            ]
        return results


class VariantUpdateSubmission:
    """
    Class to update all variant samples associated with a given
    gene list submission.
    """

    def __init__(self, filename, project, institution, vapp, validate_only=False):
        self.filename = filename
        self.project = project
        self.institution = institution
        self.vapp = vapp
        self.errors = []
        self.gene_uuids, self.bam_sample_ids = self.genes_from_file()
        self.variant_samples, self.validate_output = self.find_associated_variants()
        self.admin_vapp = self.create_admin_vapp()
        if not validate_only:
            self.post_output = self.queue_variants_for_indexing()

    def genes_from_file(self):
        """
        Extracts gene uuids and associated BAM sample IDs from input
        json file.

        :returns:
            - gene_uuids - list of gene uuids
            - bam_sample_ids - list of BAM sample IDs
        """
        gene_uuids = []
        with open(self.filename, "r") as genelist_file:
            contents = json.load(genelist_file)
        gene_uuids = contents.get("gene_uuids", [])
        bam_sample_ids = contents.get("bam_sample_ids", [])
        if not gene_uuids:
            self.errors.append("No gene uuids were found in the input file")
        return gene_uuids, bam_sample_ids

    def _search_for_variants(self, genes, project=None, add_on=None):
        """
        Helper function to search for variant samples and structural
        variant samples.

        :param genes: list of gene uuids
        :param project: str project identifier
        :param add_on: str search add-on
        :returns: list of search response items
        """
        result = []
        search_tuples = [
            ("VariantSample", "variant.genes.genes_most_severe_gene.uuid"),
            ("StructuralVariantSample", "structural_variant.transcript.csq_gene.uuid"),
        ]
        for item_type, search_term in search_tuples:
            variant_search = CommonUtils.batch_search(
                self.vapp,
                genes,
                search_term,
                batch_size=4,
                item_type=item_type,
                project=project,
                add_on=add_on,
                fields=["uuid"],
            )
            result += variant_search
        return result

    def find_associated_variants(self):
        """
        Finds associated variant samples for the genes of
        interest.

        :returns:
            - to_invalidate - list of variant sample uuids (or None)
            - validation_output - str validation message (or None)
        """
        if not self.gene_uuids:
            return None, None
        project = self.project
        genes_to_search = list(set(self.gene_uuids))
        if self.project == CGAP_CORE_PROJECT:
            project = None
        if self.bam_sample_ids:
            variant_sample_search = []
            for sample_id in self.bam_sample_ids:
                add_on = "&CALL_INFO=" + sample_id
                search = self._search_for_variants(
                    genes_to_search, project=project, add_on=add_on
                )
                variant_sample_search += search
        else:
            variant_sample_search = self._search_for_variants(
                genes_to_search, project=project
            )
        variant_sample_uuids = [item["uuid"] for item in variant_sample_search]
        to_invalidate = list(set(variant_sample_uuids))
        validation_output = "%s variant samples to update." % len(to_invalidate)
        return to_invalidate, validation_output

    def create_admin_vapp(self):
        """
        Create VirtualApp with admin permissions for posting to the
        indexing queue.

        :returns: class virtual app
        """
        app = self.vapp.app
        config = {"HTTP_ACCEPT": "application/json", "REMOTE_USER": "IMPORT"}
        admin_vapp = VirtualApp(app, config)
        return admin_vapp

    def queue_variants_for_indexing(self):
        """
        Queues all variant samples for indexing.

        :returns: str info about post (or None)
        """
        if self.errors:
            return None
        if len(self.variant_samples) == 0:
            msg = "No variant samples were posted to the indexing queue."
            return msg
        queue_body = {
            "uuids": self.variant_samples,
            "target_queue": "primary",
            "strict": True,
        }
        queue_response = self.admin_vapp.post_json("/queue_indexing", queue_body).json
        if queue_response["notification"] == "Success":
            msg = "%s variant samples successfully updated." % str(
                len(self.variant_samples)
            )
            return msg
        else:
            self.errors.append(
                "Variant samples were not properly queued for indexing."
                " Response to POSTing to /queue_indexing was: %s." % queue_response
            )
            return None


class CommonUtils:
    """
    For methods common to both submission endpoints above.
    """

    def __init__(self):
        pass

    @staticmethod
    def batch_search(
        app,
        item_list,
        search_term,
        batch_size=5,
        item_type="Gene",
        project=None,
        institution=None,
        add_on=None,
        fields=[],
    ):
        """
        Performs search requests in batches to decrease the number of
        API calls and capture all search results (as default behavior
        of vapp.get("/search/") is to return only first 25 items).

        :param app: class virtual app for search
        :param item_list: list of str items to include in search query
        :param search_term: str search term for all items in item_list
        :param batch_size: int number of items from item_list to batch
        :param item_type: str CGAP item type
        :param project: str project identifier
        :param institution: str institution identifier
        :param add_on: str search query add-on
        :param fields: list search query fields to return
        :returns: list of all items found by search
        """
        batch = []
        results = []
        flat_result = []
        search_size = 100
        add_ons = ""
        if project:
            add_ons += "&project.%40id=" + project.replace("/", "%2F")
        if institution:
            add_ons += "&institution.%40id=" + institution.replace("/", "%2F")
        if fields:
            field_string = "&field=" + "&field=".join(fields)
            add_ons += field_string
        if add_on:
            add_ons += add_on
        base_search = "/search/?type=" + item_type + "&" + search_term + "="
        for item in item_list:
            batch.append(item)
            if item == item_list[-1] or len(batch) == batch_size:
                batch_string = ("&" + search_term + "=").join(batch)
                if add_ons:
                    batch_string += add_ons
                count = 0
                new_search = True
                while new_search:
                    limit = str(search_size)
                    from_index = str(search_size * count)
                    limit_add_on = "&from=" + from_index + "&limit=" + limit
                    search_string = base_search + batch_string + limit_add_on
                    try:
                        response = app.get(search_string).json["@graph"]
                        results.append(response)
                        if len(response) == search_size:
                            count += 1
                        else:
                            new_search = False
                    except (VirtualAppError, AppError):
                        new_search = False
                batch = []
        flat_result = [x for sublist in results for x in sublist]
        return flat_result

    @staticmethod
    def is_ascii(s):
        """
        Check if characters in string are ascii.

        :param s: str
        :returns: bool if input string is ascii
        """
        if not isinstance(s, str):
            return False
        return len(s) == len(s.encode())
